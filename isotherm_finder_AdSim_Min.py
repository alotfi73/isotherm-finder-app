import numpy as np
from scipy.optimize import minimize
import os
import itertools
import tkinter as tk
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from numba import njit
import time
import threading
from tkinter import messagebox

@njit
def langmuir_single_one(p, temp, q0, b0):
        return (q0 * b0 * p) / (1 + b0 * p)
    
@njit
def langmuir_single_two(p, temp, q0, b0, b_E):
    b = b0 * np.exp(b_E/temp)
    return (q0 * b * p) / (1 + b * p)

@njit
def langmuir_single_three(p, temp, q0, q_alfa, b0, b_E):
    bp = b0 * np.exp(b_E/temp) * p
    q_max = q0 + q_alfa*temp
    return (q_max * bp) / (1 + bp)

@njit
def langmuir_single_four(p, temp, q0, q_n0, b0, b_E):
    bp = b0 * np.exp(b_E/temp) * p
    q_max = q0 / (temp**q_n0)
    return (q_max * bp) / (1 + bp)

@njit
def langmuir_single_five(p, temp, q0, q_E, b0, b_E):
    bp = b0 * np.exp(b_E/temp) * p
    q_max = q0 * np.exp(q_E/temp)
    return (q_max * bp) / (1 + bp)

@njit
def langmuir_dual_one(p, temp, q1, b1, q2, b2):
    S1 = ((q1 * b1 * p) / (1 + b1 * p)) 
    S2 = ((q2 * b2 * p) / (1 + b2 * p))
    return S1 + S2

@njit
def langmuir_dual_two(p, temp, q1, b0_1, b1_E, q2, b0_2, b2_E):
    bp1 = b0_1 * np.exp(b1_E/temp) * p
    bp2 = b0_2 * np.exp(b2_E/temp) * p
    return ((q1 * bp1) / (1 + bp2)) + ((q2 * bp1) / (1 + bp2))

@njit
def sips_LF_one(p, temp, q0, b0, n0):
    bp = b0 * p**n0
    if np.abs(bp +1) < 1e-12:
        return q0
    else:
        return (q0 * bp) / (1 + bp)

@njit
def sips_LF_two(p, temp, q0, b0, b_E, n0, alfa):
    n = n0 + alfa/temp
    bpn = b0 * np.exp(b_E/temp) * p**n
    return (q0 * bpn) / (1 + bpn)

@njit
def sips_LF_three(p, temp, q0, q_alfa, b0, b_E, n0, alfa):
    q_max = q0 + q_alfa*temp
    n = n0 + alfa/temp
    bpn = b0 * np.exp(b_E/temp) * p**n
    return (q_max * bpn) / (1 + bpn)

@njit
def sips_LF_four(p, temp, q0, q_n0, b0, b_E, n0, alfa):
    q_max = q0 / temp**q_n0
    n = n0 + alfa/temp
    bpn = b0 * np.exp(b_E/temp) * p**n
    return (q_max * bpn) / (1 + bpn)

@njit
def sips_LF_five(p, temp, q0, q_E, b0, b_E, n0, alfa):
    q_max = q0 * np.exp(q_E/temp)
    n = n0 + alfa/temp
    bpn = b0 * np.exp(b_E/temp) * p**n
    return (q_max * bpn) / (1 + bpn)

@njit
def henry_one(p, temp, param):
    return param*p

@njit
def henry_two(p, temp, q0, q_alfa):
    return (q0 + q_alfa*temp) * p

@njit
def henry_three(p, temp, q0, q_n0):
    return (q0 / temp**q_n0) * p

@njit    
def henry_four(p, temp, q0, q_E):
    return (q0 * np.exp(q_E/temp)) * p

@njit
def freundlich_one(p, temp, q0, n0):
    return q0 * (p ** n0)

@njit
def toth_one(p, temp, q0, b, n0):
    return q0*p/((b + p**n0)**(1/n0))

class isothermFinder:
    def __init__(self, data = None):
        
        self.export_name = 'isotherm_results.xlsx'
        self.R = 8.314e-3  # kJ/mol.K, m3.kPa/K.mol
        self.data =  data
        self.options = {'disp': False}
        self.results = {}

        self.cancelled = False

        self.single_dual_ratio = [20, 10, 2, 1, 0.5, 0.2, 0.1, 1e-1, 1e-2]

        self.initials = {'q':[1e-7, 1e-4, 0.1, 1],
                         'q_r':np.array([1e-7, 1e-4, 0.1, 1])/self.R,
                         'q_alfa':[-0.1, -1e-4, 1e-4, 0.1],
                        'Eq':np.array([10, 40, 80, 160])/self.R,
                        'Eb':np.array([160, 80, 40, 10])/self.R,
                        'b':[1e-10, 1e-6, 1e-3, 1e-1],
                        'k':[1e-4, 1e-2, 1, 10],
                        'n0':[1e-5, 1e-2, 1],
                        'alfa':[-100, -0.1, 0.1, 100]}
        
        self.bounds = {'q': (0, 1),
                       'q_r': (0, 1000),
                       'q_alfa':(round(-1/self.R,2), round(1/self.R,2)),
                       'Eq': (round(0.1/self.R,2), round(1000/self.R,2)),
                       'Eb': (round(0.1/self.R,2), round(1000/self.R,2)),
                       'b': (0, 100),
                       'k': (0, 100),
                       'n0': (-10, 10),
                       'alfa': (-1000,1000)}

        self.isotherms = self._define_isotherm()

        for i, iso in self.isotherms.items():
            iso["bounds"] = [self.bounds[var] for var in iso["variables"]]
            iso["initials"] = [self.initials[var] for var in iso["variables"]]

    def _define_isotherm(self):
        return {
            "Langmuir 1": {
                "func": langmuir_single_one,
                "Formula": 'IP1.IP2.P/(1+IP2.P)',
                "latex formula": r"$q = \frac{IP_1 \cdot IP_2 \cdot P}{1 + IP_2 \cdot P}$",
                "variables": ['q', 'b'],
                "units": ['kmol/kg', '-'],
                "dual": False,
                "isothermal": True
            },
            "Langmuir 2": {
                "func": langmuir_single_two,
                "Formula": 'IP1.(IP2.exp(IP3/T)).P/(1+(IP2.exp(IP3/T)).P)',
                'latex formula': r"$q = \frac{IP_1 \cdot (IP_2 \cdot e^{IP_3 / T}) \cdot P}{1 + (IP_2 \cdot e^{IP_3 / T}) \cdot P}$",
                "variables": ['q', 'b', 'Eb'],
                "units": ['kmol/kg', '-', '-'],
                "dual": False,
                "isothermal": False
            },
            "Langmuir 3": {
                "func": langmuir_single_three,
                "Formula": '(IP1+IP2.T).(IP3.exp(IP4/T)).P/(1+(IP3.exp(IP4/T)).P)',
                'latex formula': r"$q = \frac{(IP_1 + IP_2 \cdot T)(IP_3 \cdot e^{IP_4 / T}) \cdot P}{1 + (IP_3 \cdot e^{IP_4 / T}) \cdot P}$",
                "variables": ['q', 'q_alfa', 'b', 'Eb'],
                "units": ['kmol/kg', 'kmol/kg.K', '-', 'K'],
                "dual": False,
                "isothermal": False
            },
            "Langmuir 4": {
                "func": langmuir_single_four,
                "Formula": '(IP1/T^IP2).(IP3.exp(IP4/T)).P/(1+(IP3.exp(IP4/T)).P)',
                'latex formula': r"$q = \frac{(IP_1 / T^{IP_2}) \cdot IP_3 \cdot e^{IP_4 / T} \cdot P}{1 + IP_3 \cdot e^{IP_4 / T} \cdot P}$",
                "variables": ['q_r', 'n0', 'b', 'Eb'],
                "units": ['kmol.K/kg', '-', '-', 'K'],
                "dual": False,
                "isothermal": False
            },
            "Langmuir 5": {
                "func": langmuir_single_five,
                "Formula": 'IP1.exp(IP2/T).(IP3.exp(IP4/T).p)/(1+IP3.exp(IP4/T).p)',
                'latex formula': r"$q = \frac{IP_1 \cdot e^{IP_2 / T} \cdot IP_3 \cdot e^{IP_4 / T} \cdot P}{1 + IP_3 \cdot e^{IP_4 / T} \cdot P}$",
                "variables": ['q', 'Eq', 'b', 'Eb'],
                "units": ['kmol/kg', 'K', '-', 'K'],
                "dual": False,
                "isothermal": False
            },

            "Langmuir Dual 1": {
                "func": langmuir_dual_one,
                "Formula": 'IP1.IP2.P/(1+IP2.P)',
                'latex formula': r"$q = \frac{IP_1 \cdot IP_2 \cdot P}{1 + IP_2 \cdot P}$",
                "variables": ['q', 'b']*2,
                "units": ['kmol/kg', '-']*2,
                "dual": True,
                "isothermal": True,
                "single iso": 'Langmuir 1'
            },
            "Langmuir Dual 2": {
                "func": langmuir_dual_two,
                "Formula": 'IP1.(IP2.exp(IP3/T)).P/(1+(IP2.exp(IP3/T)).P)',
                'latex formula': r"$q = \frac{IP_1 \cdot IP_2 \cdot e^{IP_3 / T} \cdot P}{1 + IP_2 \cdot e^{IP_3 / T} \cdot P}$",
                "variables": ['q', 'b', 'Eb']*2,
                "units": ['kmol/kg', '-', 'K']*2,
                "dual": True,
                "isothermal": False,
                "single iso": 'Langmuir 2'
            },

            "Sips LF 1": {
                "func": sips_LF_one,
                "Formula": 'IP1.IP2.P^IP3/(1+IP2.P^IP3)',
                'latex formula': r"$q = \frac{IP_1 \cdot IP_2 \cdot P^{IP_3}}{1 + IP_2 \cdot P^{IP_3}}$",
                "variables": ['q', 'b', 'n0'],
                "units": ['kmol/kg', '-', '-'],
                "dual": False,
                "isothermal": True
            },
            "Sips LF 2": {
                "func": sips_LF_two,
                "Formula": 'IP1.IP2.exp(IP3/T).P^(IP4+IP5/T)/(1+IP2.exp(IP3/T).P^(IP4+IP5/T))',
                'latex formula': r"$q = \frac{IP_1 \cdot IP_2 \cdot e^{IP_3 / T} \cdot P^{(IP_4 + IP_5 / T)}}{1 + IP_2 \cdot e^{IP_3 / T} \cdot P^{(IP_4 + IP_5 / T)}}$",
                "variables": ['q', 'b', 'Eb', 'n0', 'alfa'],
                "units": ['kmol/kg', '-', 'K', '-', 'K'],
                "dual": False,
                "isothermal": False
            },
            "Sips LF 3": {
                "func": sips_LF_three,
                "Formula": '(IP1+IP2.T).IP3.exp(IP4/T).P^(IP5+IP6/T)/(1+IP3exp(IP4/T).P^(IP5+IP6/T))',
                'latex formula': r"$q = \frac{(IP_1 + IP_2 \cdot T) \cdot IP_3 \cdot e^{IP_4 / T} \cdot P^{(IP_5 + IP_6 / T)}}{1 + IP_3 \cdot e^{IP_4 / T} \cdot P^{(IP_5 + IP_6 / T)}}$",
                "variables": ['q', 'q_alfa', 'b', 'Eb', 'n0', 'alfa'],
                "units": ['kmol/kg', 'kmol/kg.K', '-', 'K', '-', 'K'],
                "dual": False,
                "isothermal": False
            },
            
            "Sips LF 4": {
                "func": sips_LF_four,
                "Formula": '(IP1/T^IP2).IP3.exp(IP4/T).P^(IP5+IP6/T)/(1+IP3exp(IP4/T).P^(IP5+IP6/T))',
                'latex formula': r"$q = \frac{(IP_1 / T^{IP_2}) \cdot IP_3 \cdot e^{IP_4 / T} \cdot P^{(IP_5 + IP_6 / T)}}{1 + IP_3 \cdot e^{IP_4 / T} \cdot P^{(IP_5 + IP_6 / T)}}$",
                "variables": ['q_r', 'n0', 'b', 'Eb', 'n0', 'alfa'],
                "units": ['kmol/kg', '-', '-', 'K', '-', 'K'],
                "dual": False,
                "isothermal": False
            },
            "Sips LF 5": {
                "func": sips_LF_five,
                "Formula": 'IP1.exp(IP2/T).IP3.exp(IP4/T).P^(IP5+IP6/T)/(1+IP23exp(IP4/T).P^(IP5+IP6/T))',
                'latex formula': r"$q = \frac{IP_1 \cdot e^{IP_2 / T} \cdot IP_3 \cdot e^{IP_4 / T} \cdot P^{IP_5 + IP_6 / T}}{1 + IP_{23} \cdot e^{IP_4 / T} \cdot P^{IP_5 + IP_6 / T}}$",
                "variables": ['q', 'Eq', 'b', 'Eb', 'n0', 'alfa'],
                "units": ['kmol/kg', 'K', '-', 'K', '-', 'K'],
                "dual": False,
                "isothermal": False
            },
            
            "Henry 1": {
                "func": henry_one,
                "Formula": 'IP1.P',
                'latex formula': r"$q = IP_1 \cdot P$",
                "variables": ['q'],
                "units": ['kmol/kg'],
                "dual": False,
                "isothermal": False
            },
            "Henry 2": {
                "func": henry_two,
                "Formula": '(IP1+IP2.T).P',
                'latex formula': r"$q = (IP_1 + IP_2 \cdot T) \cdot P$",
                "variables": ['q', 'q_alfa'],
                "units": ['kmol/kg', 'kmol/kg.K'],
                "dual": False,
                "isothermal": False
            },
            "Henry 3": {
                "func": henry_three,
                "Formula": '(IP1/T^IP2).P',
                'latex formula': r"$q = (IP_1 / T^{IP_2}) \cdot P$",
                "variables": ['q_r', 'n0'],
                "units": ['kmol.T/kg', '-'],
                "dual": False,
                "isothermal": False
            },
            "Henry 4": {
                "func": henry_four,
                "Formula": '(IP1.exp(IP2/T).P',
                'latex formula': r"$q = IP_1 \cdot e^{IP_2 / T} \cdot P$",
                "variables": ['q', 'Eq'],
                "units": ['kmol.T/kg', 'K'],
                "dual": False,
                "isothermal": False
            },

            "Freundlich 1": {
                "func": freundlich_one,
                "Formula": 'IP1.P^IP2',
                'latex formula': r"$q = IP_1 \cdot P^{IP_2}$",
                "variables": ['q', 'n0'],
                "units": ['kmol.T/kg', '-'],
                "dual": False,
                "isothermal": False
            },

            "Toth 1": {
                "func": toth_one,
                "Formula": 'IP1.P/(IP2+P^IP3)^1/IP3',
                'latex formula': r"$q = \frac{IP_1 \cdot P}{(IP_2 + P^{IP_3})^{1 / IP_3}}$",
                "variables": ['q', 'b', 'n0'],
                "units": ['kmol.T/kg', '-', '-'],
                "dual": False,
                "isothermal": False
            }
        }
 
    def mean_squared_error(self, y_true, y_pred):
        y_true = np.asarray(y_true)
        y_pred = np.asarray(y_pred)
        return np.mean(((y_true - y_pred)*1000) ** 2)
    
    def r2_score(self, y_true, y_pred):
        y_true = np.asarray(y_true)
        y_pred = np.asarray(y_pred)
        ss_res = np.sum((y_true - y_pred) ** 2)
        ss_tot = np.sum((y_true - np.mean(y_true)) ** 2)
        return 1 - ss_res / ss_tot
    
    def report_error(self, exp, cal):
        rmse = np.sqrt(self.mean_squared_error(exp, cal))
        r2 = self.r2_score(exp, cal)
        return [rmse *100, r2]
    
    def fit_model(self, model_func, initial_guess, p_exp, q_exp, temp, bounds=None):
        
        
        best_res = None
        lowest_cost = np.inf
        
        for i_g in initial_guess:
            res = minimize(
                self.min_func, i_g, args=(model_func, temp, p_exp, q_exp), 
                method=self.opt_method, bounds=bounds, options=self.options
            )
            if res.success and res.fun < lowest_cost:
                best_res = res
                lowest_cost = res.fun
        
        
        #with parallel_backend('loky'):
        #    results = Parallel(n_jobs=-1)(delayed(minimize)(
        #       self.min_func, i_g, args=(model_func, temp, p_exp, q_exp), 
        #        method=self.opt_method, bounds=bounds, options=self.options
        #    ) for i_g in initial_guess)

        #best_res = min(results, key=lambda x: x.fun if x.success else np.inf)

        for _ in range(self.opt_accuracy):
            opt_res = minimize(self.min_func, best_res.x, args=(model_func, temp, p_exp, q_exp), 
                               method=self.opt_method, bounds=bounds,options=self.options)
            if opt_res.fun < best_res.fun:
                best_res = opt_res
            else:
                break
        
        return best_res.x, best_res.fun

    def set_all_initial(self, isotherm):
        # Get the list of values for each variable from the initials dictionary
        value_lists = isotherm["initials"]
        # Generate all possible combinations
        return list(itertools.product(*value_lists))
    
    def set_initial_from_single(self, single_iso):
        param = self.results[single_iso]["params"]
        return [np.concatenate((param, i * param)) for i in self.single_dual_ratio]

    def min_func(self, params, model_func, temp, pres, q_exp):
        if self.cancelled:
            raise Exception("Optimization Cancelled")
        
        q_calc = [model_func(p, t, *params) for p,t in zip(pres,temp)]
        return np.average([np.mean(((q_e - q_c)*1000) ** 2) for q_e,q_c in zip(q_exp,q_calc)])*100

    def set_cal_data(self, model_func, params, p_exp, q_exp, temp):
        q_cal = [model_func(p,t, *params) for p,t in zip(p_exp,temp)]
        errors = [np.mean(col) for col in zip(*[self.report_error(y0, y1) for y0, y1 in zip(q_exp, q_cal)])]
        return errors, q_cal
    
    def plot_results(self, address=None):
        import matplotlib.pyplot as plt

        for iso_name, res in self.results.items():
            temp, p_exp, q_exp, result_data = self.data['temp']-273.15, self.data['p_exp'], self.data['q_exp'], res

            if iso_name == 'best iso':
                return None

            plt.figure(figsize=(8, 6))
            for temp, p, q, q_fit in zip(temp, p_exp, q_exp, result_data['q_cal']):
                plt.scatter(p, q*1000, label=f"Exp: {round(temp,2)}°C", alpha=0.7)
                plt.plot(p, q_fit*1000, label=f"Fit: {round(temp,2)}°C", linestyle="--")
            plt.title(f"{iso_name} isotherm with {round(result_data['RMSE'],2)}% error")
            plt.xlabel("Pressure (bar)")
            plt.ylabel("Adsorption (mol/g)")
            plt.legend()
            plt.show(block=False)
            if address:
                plot_path = os.path.join(address, f'{iso_name}.png')
                plt.savefig(plot_path)
                plt.close()
            
    def plot_best_fit(self, address = None):
        import matplotlib.pyplot as plt

        model = self.results['best iso']
        plot_title = f'Best isotherm fit ({self.results["best iso"]})'
        temp, p_exp, q_exp, q_cal = self.data['temp']-273.15, self.data['p_exp'], self.data['q_exp'], self.results[model]['q_cal']

        plt.figure(figsize=(8, 6))
        for temp, p, q, q_fit in zip(temp, p_exp, q_exp, q_cal):
            plt.scatter(p, q*1000, label=f"Exp: {round(temp,2)}°C", alpha=0.7)
            plt.plot(p, q_fit*1000, label=f"Fit: {round(temp,2)}°C", linestyle="--")
        plt.title(plot_title)
        plt.xlabel("Pressure (bar)")
        plt.ylabel("Adsorption (mol/g)")
        plt.legend()
        plt.show(block=False)
        if address:
            plot_path = os.path.join(address, f'{plot_title}.png')
            plt.savefig(plot_path)
            plt.close()

    def export_results(self, gas, address):
        import pandas as pd

        sel_keys_iso, sel_keys_res = ["func", "Formula"], ["RMSE", "R2_errors"]
        path = os.path.join(address, self.export_name)
        with pd.ExcelWriter(path, engine='xlsxwriter') as writer:
            best_iso = self.results['best iso']
            iso, res = self.isotherms[best_iso], self.results[best_iso]
            result = {'best iso': [best_iso] + 
                    [item for k in sel_keys_iso if k in iso for item in (k, iso[k])] +
                    [item for k in sel_keys_res if k in res for item in (k, res[k])]}
            
            params = self.results[best_iso]["params"]
            result.update({"params": [f'IP{i+1}' for i in range(len(params))],
                        "Values": params})

            idx = 1
            for key in self.results.keys():
                if key != best_iso and key != 'best iso':
                    iso, res = self.isotherms[key], self.results[key]
                    result.update({f'{idx+1}': [key] + 
                                [item for k in sel_keys_iso if k in iso for item in (k, iso[k])] +
                                [item for k in sel_keys_res if k in res for item in (k, res[k])]})
                    
                    params = self.results[key]["params"]
                    result.update({f"params {idx+1}": [f'IP{i+1}' for i in range(len(params))],
                        f"Values {idx+1}": params})
                    idx += 1

            result = {k: v if isinstance(v, list) else list(v) if isinstance(v, np.ndarray) else [v] for k, v in result.items()}
            
            max_len = max(map(len, result.values()))
            result = {k: v + [np.nan] * (max_len - len(v)) for k, v in result.items()}
        
            df = pd.DataFrame(result)
            df.to_excel(writer, sheet_name=gas, index=False)
        self.polish_excel_format(path)
    
    def polish_excel_format(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        
        # Define styles
        fill_gray = PatternFill(start_color="D9D9D9", fill_type="solid")
        fill_blue = PatternFill(start_color="BDD7EE", fill_type="solid")
        fill_yellow = PatternFill(start_color="FFFFCC", fill_type="solid")
        fill_green = PatternFill(start_color="CCFFCC", fill_type="solid")
        fill_red = PatternFill(start_color="FFCCCC", fill_type="solid")
        fill_light_yellow = PatternFill(start_color="FFF2CC", fill_type="solid")
        
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )

        for sheet in wb.worksheets:
            max_col = sheet.max_column
            max_row = sheet.max_row

            for col_start in range(1, max_col + 1, 3):
                col_end = min(col_start + 2, max_col)

                # Set header fill color (first row)
                header_fill = fill_gray if ((col_start - 1) // 3) % 2 == 0 else fill_blue
                for col in range(col_start, col_end + 1):
                    cell = sheet.cell(row=1, column=col)
                    cell.fill = header_fill

                # Set alternating row fill colors for every 3-column group
                for row in range(2, max_row + 1):
                    for col in range(col_start, col_end + 1):
                        cell = sheet.cell(row=row, column=col)
                        if (row % 2 == 0) and (col - col_start == 0):  # First col: row 2,4,... → yellow
                            cell.fill = fill_yellow
                        elif (row % 2 == 1) and (col - col_start == 0):  # First col: row 3,5,... → green
                            cell.fill = fill_green
                        elif (col - col_start == 1):  # Second col: param names
                            if sheet.cell(row=row, column=col).value is not None:
                                cell.fill = fill_red
                        elif (col - col_start == 2):  # Third col: param values
                            if sheet.cell(row=row, column=col).value is not None:
                                cell.fill = fill_light_yellow

                # Apply thick border around 3-column group
                for row in range(1, max_row + 1):
                    for col in range(col_start, col_end + 1):
                        cell = sheet.cell(row=row, column=col)
                        border_sides = {
                            'left': Side(style='thick') if col == col_start else Side(style='thin'),
                            'right': Side(style='thick') if col == col_end else Side(style='thin'),
                            'top': Side(style='thick') if row == 1 else Side(style='thin'),
                            'bottom': Side(style='thick') if row == max_row else Side(style='thin'),
                        }
                        cell.border = Border(**border_sides)

        wb.save(file_path)

    def check_dual_single(self):
        self.added_iso = []
        for i, iso in enumerate(self.selected_isothemrs):
            if self.isotherms[iso]["dual"]:
                single = self.isotherms[iso]["single iso"]
                if not single in self.selected_isothemrs:
                    self.added_iso.append(single)
                    self.selected_isothemrs.insert(i, single) 

    
    def find_isotherm(self, data, selected_isothemrs = [], opt_method = 'Nelder-Mead', accuracy = 5):

        self.selected_isothemrs = selected_isothemrs
        self.data = data
        self.opt_method = opt_method
        self.opt_accuracy = accuracy

        self.check_dual_single()

        p_exp, q_exp, temp = self.data['p_exp'], self.data['q_exp'], self.data['temp']

        self.results= {}

        count = 1

        for name in self.selected_isothemrs:
            t0 = time.time()
            details = self.isotherms[name]

            self.toast = tk.Toplevel()
            self.toast.title("Fitting Status")
            self.toast.geometry("900x200")

            tk.Label(self.toast, text=f"isotherm fitting {count} of {len(self.selected_isothemrs)}: {name} started...", 
                     padx=20, pady=10, font=("Helvetica", 16, "bold")).pack()

            cancel_btn = tk.Button(self.toast, text="Cancel", command=self.cancel)
            cancel_btn.pack(pady=10)

            self.toast.update_idletasks()
            self.toast.update()
            self.toast.protocol("WM_DELETE_WINDOW", self.cancel)

            list_guess = self.set_initial_from_single(details["single iso"]) if details["dual"] else self.set_all_initial(details)

            try:
                params, opt_errors = self.fit_model(
                    details["func"],
                    list_guess,
                    p_exp, q_exp, temp,
                    bounds=details["bounds"],
                )
            except Exception as e:
                if str(e) == "Optimization Cancelled":
                    messagebox.showwarning("Cancellation", "Optimization was cancelled")
                    return False

            errors, q_cal = self.set_cal_data(details["func"], params, p_exp, q_exp, temp)

            self.results[name] = {"params": params, "RMSE": errors[0], "R2_errors": errors[1], 
                                         "q_cal": q_cal, "opt_errors": opt_errors}
            
            self.toast.destroy()
            count += 1

            print(f'time for {name}: {time.time()-t0}')

        [self.results.pop(key, None) for key in self.added_iso]

        best_iso_name = min(self.results.items(), key=lambda x: x[1]["opt_errors"])[0]

        self.results['best iso'] = best_iso_name

    def cancel(self):
        self.cancelled = True
        if hasattr(self, 'toast'):
            self.toast.destroy()

    def start_fitting_thread(self, callback, *args):
        self.thread = threading.Thread(target=self._thread_wrapper, args=(callback, *args))
        self.thread.start()

    def _thread_wrapper(self, callback, *args):
        self.find_isotherm(*args)
        self.done_callback = callback

